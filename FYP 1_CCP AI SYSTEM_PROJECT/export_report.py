import argparse
import base64
import json
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.worksheet.views import Selection
from openpyxl.utils.units import pixels_to_EMU


ZONE_AREAS = {
    "1": "Underhood", "2": "Underhood", "3": "Underhood", "4": "Underhood", "5": "Underhood", "6": "Underhood",
    "11": "Underbody", "12": "Underbody", "13": "Underbody", "14": "Underbody", "15": "Underbody", "16": "Underbody", "17": "Underbody",
    "20": "Under seats / instrument panel", "21": "Under seats / instrument panel", "22A": "Under seats / instrument panel",
    "22B": "Under seats / instrument panel", "23A": "Under seats / instrument panel", "23B": "Under seats / instrument panel",
    "28": "Under seats / instrument panel",
    "24": "Trunk / wheel arches / front grille", "25": "Trunk / wheel arches / front grille",
    "26": "Trunk / wheel arches / front grille", "27": "Trunk / wheel arches / front grille",
    "29": "Trunk / wheel arches / front grille",
}


def text(value):
    return "" if value is None else str(value)


def index_number(row):
    zone = text(row.get("zone")).strip()
    return f"Z{zone}" if zone else text(row.get("checkpoint")).strip()


def area_name(row):
    zone = text(row.get("zone")).strip()
    return ZONE_AREAS.get(zone, f"Zone {zone}" if zone else "")


def default_detail(row):
    return text(row.get("observation")).strip()


def status_from_grade(row):
    grade = text(row.get("grade")).strip().upper()
    if grade in {"NO GRADE", "NOGRADE", "NO-GRADE", "NA", "N/A"}:
        return "NA"
    if grade in {"S", "A", "B", "C"}:
        return "OK"
    return text(row.get("status")).strip()


def has_photo_data(row):
    return bool(photo_data_list(row))


def photo_data_list(row):
    values = row.get("photoDataList")
    if isinstance(values, list):
        photos = [text(value) for value in values if text(value).startswith("data:image/")]
        if photos:
            return photos[:4]
    single = text(row.get("photoData"))
    return [single] if single.startswith("data:image/") else []


def defect_area_image_data(row):
    value = text(row.get("defectAreaImageData"))
    return value if value.startswith("data:image/") else ""


def row_values(row, fallback_no):
    zone = text(row.get("zone")).strip()
    return [
        row.get("concernNo") or fallback_no,
        index_number(row),
        text(row.get("vehicle")),
        area_name(row),
        text(row.get("powertrain")),
        text(row.get("system")),
        text(row.get("routingType")),
        default_detail(row),
        text(row.get("clearanceMiniAcceptable")),
        "" if defect_area_image_data(row) else text(row.get("defectArea")),
        "" if has_photo_data(row) else text(row.get("photo")),
        text(row.get("alignedComment")),
        status_from_grade(row),
        text(row.get("grade")),
    ]


def find_session_sheet(workbook):
    for sheet in workbook.worksheets:
        if sheet.title.strip().upper().startswith("SESSION"):
            return sheet
    raise ValueError("No SESSION sheet found in the template.")


def find_header_row(sheet):
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        value = text(sheet.cell(row_index, 1).value).strip().upper()
        if value == "SL NO":
            return row_index
    raise ValueError("No report header row found in the SESSION sheet.")


def first_non_empty(rows, key):
    for row in rows:
        value = text(row.get(key)).strip()
        if value:
            return value
    return ""


def fill_report_metadata(sheet, rows, header_row):
    sorted_rows = sorted(rows, key=lambda item: int(item.get("concernNo") or 0))
    inspector = first_non_empty(sorted_rows, "inspector")
    inspection_date = first_non_empty(sorted_rows, "date") or datetime.now().strftime("%Y-%m-%d")
    export_time = datetime.now().strftime("%H:%M")
    name_value = f"Name : {inspector}".strip()
    date_time_value = f"Date & Time : {inspection_date} {export_time}".strip()

    name_cell = None
    date_time_cell = None
    for row in sheet.iter_rows(min_row=1, max_row=max(1, header_row - 1)):
        for cell in row:
            value = text(cell.value).strip().lower()
            if "name" in value and name_cell is None:
                name_cell = cell
            if "date" in value and "time" in value and date_time_cell is None:
                date_time_cell = cell

    if name_cell is None and date_time_cell is None:
        sheet.insert_rows(1)
        header_row += 1
        sheet["A1"].value = name_value
        sheet["K1"].value = date_time_value
        try:
            sheet.merge_cells("A1:I1")
            sheet.merge_cells("K1:L1")
        except ValueError:
            pass
        return header_row

    if name_cell is not None:
        name_cell.value = name_value
    if date_time_cell is not None:
        date_time_cell.value = date_time_value
    return header_row


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def clear_template_images(sheet):
    # The report template can contain example/placeholder images in blank rows.
    # Remove them during export so only current saved inspection images appear.
    if hasattr(sheet, "_images"):
        sheet._images = []


def reset_report_open_view(workbook, sheet, first_data_row):
    workbook.active = workbook.index(sheet)
    first_data_cell = f"A{first_data_row}"
    sheet.freeze_panes = first_data_cell
    sheet.sheet_view.topLeftCell = "A1"
    if sheet.sheet_view.pane:
        sheet.sheet_view.pane.topLeftCell = first_data_cell
        sheet.sheet_view.pane.activePane = "bottomLeft"
    sheet.sheet_view.selection = [Selection(pane="bottomLeft", activeCell=first_data_cell, sqref=first_data_cell)]


def safe_sheet_title(label):
    cleaned = "".join(ch for ch in text(label).upper() if ch not in r'[]:*?/\\').strip()
    return (cleaned or "SESSION")[:31]


def add_data_url_image(sheet, row_index, data_url, excel_col_zero_based, fallback_col_one_based, cell_width, cell_height, margin=8):
    if not data_url or "," not in data_url:
        return False
    try:
        raw = base64.b64decode(data_url.split(",", 1)[1])
        image = ExcelImage(BytesIO(raw))
        scale = min(1, (cell_width - margin * 2) / image.width, (cell_height - margin * 2) / image.height)
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
        x_offset = margin + max(0, int((cell_width - margin * 2 - image.width) / 2))
        y_offset = margin + max(0, int((cell_height - margin * 2 - image.height) / 2))
        marker = AnchorMarker(col=excel_col_zero_based, colOff=pixels_to_EMU(x_offset), row=row_index - 1, rowOff=pixels_to_EMU(y_offset))
        image.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(image.width), pixels_to_EMU(image.height)))
        sheet.add_image(image)
        return True
    except Exception:
        sheet.cell(row_index, fallback_col_one_based).value = "Image could not be inserted"
        return False


def add_defect_area_image(sheet, row_index, row):
    image_data = defect_area_image_data(row)
    if not image_data:
        return
    sheet.row_dimensions[row_index].height = max(sheet.row_dimensions[row_index].height or 0, 234.9)
    add_data_url_image(sheet, row_index, image_data, 9, 10, 315, 300, margin=4)


def add_photo(sheet, row_index, row):
    photos = photo_data_list(row)
    if not photos:
        return
    sheet.row_dimensions[row_index].height = max(sheet.row_dimensions[row_index].height or 0, 234.9)
    cell_width = 420
    cell_height = 300
    margin = 8
    gap = 8
    grid_cols = 1 if len(photos) == 1 else 2
    grid_rows = 1 if len(photos) <= 2 else 2
    slot_width = int((cell_width - (margin * 2) - (gap * (grid_cols - 1))) / grid_cols)
    slot_height = int((cell_height - (margin * 2) - (gap * (grid_rows - 1))) / grid_rows)
    for index, photo_data in enumerate(photos[:4]):
        if "," not in photo_data:
            continue
        try:
            raw = base64.b64decode(photo_data.split(",", 1)[1])
            image = ExcelImage(BytesIO(raw))
            col_index = index % grid_cols
            row_index_in_cell = index // grid_cols
            scale = min(1, slot_width / image.width, slot_height / image.height)
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
            x_offset = margin + col_index * (slot_width + gap) + max(0, int((slot_width - image.width) / 2))
            y_offset = margin + row_index_in_cell * (slot_height + gap) + max(0, int((slot_height - image.height) / 2))
            marker = AnchorMarker(col=10, colOff=pixels_to_EMU(x_offset), row=row_index - 1, rowOff=pixels_to_EMU(y_offset))
            image.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(image.width), pixels_to_EMU(image.height)))
            sheet.add_image(image)
        except Exception:
            sheet.cell(row_index, 11).value = text(row.get("photo")) or "Photo could not be inserted"
            return


def build_report(template_path, input_path, output_path):
    payload = json.loads(Path(input_path).read_text(encoding="utf-8-sig"))
    rows = payload.get("rows") or []
    session_label = payload.get("sessionLabel") or "Session"

    workbook = load_workbook(template_path, keep_vba=True)
    sheet = find_session_sheet(workbook)
    clear_template_images(sheet)
    sheet.title = safe_sheet_title(session_label)
    header_row = find_header_row(sheet)
    header_row = fill_report_metadata(sheet, rows, header_row)
    sheet.cell(header_row, 9).value = "Clearance mini acceptable + put the source of the value in DT"

    max_columns = 14
    template_row = header_row + 1
    template_height = sheet.row_dimensions[template_row].height

    last_row = max(sheet.max_row, template_row + len(rows) - 1)
    for row_index in range(template_row, last_row + 1):
        if template_height:
            sheet.row_dimensions[row_index].height = template_height
        for column_index in range(1, max_columns + 1):
            cell = sheet.cell(row_index, column_index)
            cell.value = None
            copy_cell_style(sheet.cell(template_row, column_index), cell)

    sorted_rows = sorted(rows, key=lambda item: int(item.get("concernNo") or 0))
    for fallback_no, record in enumerate(sorted_rows, start=1):
        row_index = template_row + fallback_no - 1
        for column_index, value in enumerate(row_values(record, fallback_no), start=1):
            sheet.cell(row_index, column_index).value = value
        add_defect_area_image(sheet, row_index, record)
        add_photo(sheet, row_index, record)

    reset_report_open_view(workbook, sheet, template_row)
    workbook.save(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fill the CCP report XLSM template with one session of saved inspections.")
    parser.add_argument("--template", required=True)
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    build_report(args.template, args.input, args.output)
